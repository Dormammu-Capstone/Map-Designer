# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


# def print_hi(name):
# Use a breakpoint in the code line below to debug your script.
# print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
# if __name__ == '__main__':
#    print_hi('PyCharm')*/

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
import os
import sys
import random

from PySide6 import *
from PySide6.QtCore import QFileInfo
from PySide6.QtGui import Qt, QPixmap, QIcon, QBrush, QFont, QColor
from PySide6.QtUiTools import loadUiType
import random

import xlsxwriter
import openpyxl
from PyQt5.uic.properties import QtGui

import pymysql
from PySide6.QtWidgets import QMainWindow, QDialog, QWidget, QFileDialog, QTableWidget, QVBoxLayout, QGridLayout, \
    QPushButton, QTableWidgetItem, QApplication, QLabel

from pymysql.constants import CLIENT

current_path = os.path.dirname(os.path.abspath(__file__))
r_path = os.path.join(current_path, "image", "r_arrow.png")
u_path = os.path.join(current_path, "image", "u_arrow.png")
l_path = os.path.join(current_path, "image", "l_arrow.png")
d_path = os.path.join(current_path, "image", "d_arrow.png")
d_l_path = os.path.join(current_path, "image", "d_l_arrow.png")
d_r_path = os.path.join(current_path, "image", "d_r_arrow.png")
r_l_path = os.path.join(current_path, "image", "r_l_arrow.png")
u_d_path = os.path.join(current_path, "image", "u_d_arrow.png")
u_l_path = os.path.join(current_path, "image", "u_l_arrow.png")
u_r_path = os.path.join(current_path, "image", "u_r_arrow.png")
zero_path = os.path.join(current_path, "image", "zero.png")

conn = None
cur = None

conn = pymysql.connect(host='127.0.0.1', port=3306, user='root', password='1290', db='lghpdb', charset='utf8',
                       client_flag=CLIENT.MULTI_STATEMENTS, autocommit=True)
cur = conn.cursor()


def resource_path(relative_path):
    base_path = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


# 이미지랑 텍스트 넣기


def dir_img(path, text):
    pixmap = QPixmap(path)
    pixmap.scaled(1, 1)
    icon = QIcon(pixmap)
    item = QTableWidgetItem()
    item.setText(text)
    item.setIcon(icon)
    font = QFont()
    font.setPointSize(1)
    item.setFont(font)
    item.setForeground(QBrush(Qt.transparent))
    item.setTextAlignment(Qt.AlignCenter)
    return item
def but_img(path):
    pixmap = QPixmap(path)
    pixmap.scaled(1, 1)
    icon = QIcon(pixmap)
    but = QPushButton()
    #but.setText(text)
    but.setIcon(icon)
    #font = QFont()
    #font.setPointSize(1)
    #item.setFont(font)
    #item.setForeground(QBrush(Qt.transparent))
    #item.setTextAlignment(Qt.AlignCenter)
    return but


# 1.homePage.ui
form = resource_path('homePage.ui')  # 여기에 ui파일명 입력
# form_class = uic.loadUiType(form)[0]
form_class = loadUiType(form)[0]
# 2.setGrid.ui
form_second = resource_path('setGrid.ui')
# form_secondwindow = uic.loadUiType(form_second)[0]
form_secondwindow = loadUiType(form_second)[0]
# 3.setAttribute.ui
form_third = resource_path('setAttribute.ui')
# form_thirdwindow = uic.loadUiType(form_third)[0]
form_thirdwindow = loadUiType(form_third)[0]
# 4.createMap.ui
form_fourth = resource_path('createMap.ui')
# form_fourthwindow = uic.loadUiType(form_fourth)[0]
form_fourthwindow = loadUiType(form_fourth)[0]
# 5.viewFile.ui
form_fifth = resource_path('viewFile.ui')
# form_fifthwindow = uic.loadUiType(form_fifth)[0]
form_fifthwindow = loadUiType(form_fifth)[0]
# 6.sixthFile.ui
form_sixth = resource_path('editFile.ui')
# form_sixthwindow = uic.loadUiType(form_sixth)[0]
form_sixthwindow = loadUiType(form_sixth)[0]


# 1.homePage.ui


class WindowClass(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("맵 디자인")
        self.setFixedSize(1000, 550)
        self.createFile.clicked.connect(
            self.btn_createfile_to_setgrid)  # createFile button 클릭
        self.openFile.clicked.connect(self.btn_fileLoad)  # openFile button 클릭

    # 여기에 시그널-슬롯 연결 설정 및 함수 설정.
    # -createFile button 함수: setGrid.ui로 창전환
    def btn_createfile_to_setgrid(self):
        self.hide()
        self.second = secondwindow()
        self.second.exec()
        self.show()

    # -openFile button 함수: 파일선택창
    def btn_fileLoad(self):
        # 미리보기ui연결 수정클릭->수정페이지
        self.hide()
        self.fifth = fifthwindow()
        self.fifth.exec()
        self.show()  # homepage로 돌아감


# 5. viewFile.ui
class fifthwindow(QDialog, QWidget, form_fifthwindow):
    def __init__(self, parent=None):
        global row, col, file_name
        super(fifthwindow, self).__init__()
        # self.initUi()
        self.setupUi(self)
        self.setWindowTitle("맵 미리보기")
        self.show()  # 파일선택후 창이 앞으로 띄워지게 하기위해 위에 위치
        self.setFixedSize(1000, 550)
        file = QFileDialog.getOpenFileName(
            self, '', '', 'xlsx파일 (*.xlsx);; All File(*)')  # !!저장파일 타입 정해지면, 확장자에 추가
        global filename  # 선언, 할당 분리
        filename = file[0]
        load_xlsx = openpyxl.load_workbook(file[0], data_only=True)
        load_sheet = load_xlsx['NewSheet1']
        self.table = QTableWidget(parent)

        # 파일 이름으로 db에서 해당 정보 연결
        file_name = QFileInfo(file[0]).baseName()

        vbox = QVBoxLayout()
        vbox.addWidget(self.table)
        grid = QGridLayout()
        vbox.addLayout(grid)
        edit = QPushButton("수정")
        grid.addWidget(edit, 0, 0)
        self.setLayout(vbox)
        self.setGeometry(200, 200, 400, 500)
        sql = "SELECT GridSizeX FROM grid " + "WHERE Grid_ID = %s;"
        cur.execute(sql, [str(file_name)])
        file_col = cur.fetchone()
        sql = "SELECT GridSizeY FROM grid " + "WHERE Grid_ID = %s;"
        cur.execute(sql, [str(file_name)])
        file_row = cur.fetchone()
        row = int(file_row[0])
        col = int(file_col[0])

        self.table.setColumnCount(col)
        self.table.setRowCount(row)
        # 반드시 item 생성해야 셀 색상 변경가능
        for i in range(row):
            for j in range(col):
                self.table.setItem(i, j, QTableWidgetItem())
        self.table.resizeColumnsToContents()
        self.table.resizeRowsToContents()
        # load_excel은 1부터, table은 0부터

        cnum = 1
        for i in range(1, row + 1):
            for j in range(1, col + 1):
                print(load_sheet.cell(i, j).fill.start_color.index)
                cell_num = str(file_name) + '_c' + str(cnum).zfill(4)
                cnum = cnum + 1
                sql = "SELECT * FROM cell " + "WHERE Cell_ID = %s;"
                cur.execute(sql, [cell_num])
                cellinfo = cur.fetchone()
                # n:6 s:7 w:8 e:9
                if cellinfo[6] == 1:
                    item = dir_img(u_path, "↑")
                    self.table.setItem(i - 1, j - 1, item)
                    # self.table.item(i - 1, j - 1).setText("↑")
                elif cellinfo[7] == 1:
                    item = dir_img(d_path, "↓")
                    self.table.setItem(i - 1, j - 1, item)
                    # self.table.item(i - 1, j - 1).setText("↓")
                elif cellinfo[8] == 1:
                    item = dir_img(l_path, "←")
                    self.table.setItem(i - 1, j - 1, item)
                    # self.table.item(i - 1, j - 1).setText("←")
                elif cellinfo[9] == 1:
                    item = dir_img(r_path, "→")
                    self.table.setItem(i - 1, j - 1, item)
                    # self.table.item(i - 1, j - 1).setText("→")
                """elif load_sheet.cell(i, j).text()=="←↓": #db로 바꾸기
                    item = dir_img(d_l_path, "←↓")
                    self.table.setItem(i - 1, j - 1, item)
                elif load_sheet.cell(i, j).text()=="↓→": #db로 바꾸기
                    item = dir_img(d_r_path, "↓→")
                    self.table.setItem(i - 1, j - 1, item)
                elif load_sheet.cell(i, j).text()=="↔": #db로 바꾸기
                    item = dir_img(r_l_path, "↔")
                    self.table.setItem(i - 1, j - 1, item)
                elif load_sheet.cell(i, j).text()=="↕": #db로 바꾸기
                    item = dir_img(u_d_path, "↕")
                    self.table.setItem(i - 1, j - 1, item)
                elif load_sheet.cell(i, j).text()=="←↑": #db로 바꾸기
                    item = dir_img(u_l_path, "←↑")
                    self.table.setItem(i - 1, j - 1, item)
                elif load_sheet.cell(i, j).text()=="↑→": #db로 바꾸기
                    item = dir_img(u_r_path, "↑→")
                    self.table.setItem(i - 1, j - 1, item)"""
                if load_sheet.cell(i, j).fill.start_color.index == 'FFFFFF00':
                    self.table.item(i - 1, j - 1).setBackground(Qt.yellow)
                    self.table.item(i - 1, j - 1).setForeground(Qt.black)
                elif load_sheet.cell(i, j).fill.start_color.index == 'FF0000FF':
                    self.table.item(i - 1, j - 1).setBackground(Qt.darkBlue)
                    self.table.item(i - 1, j - 1).setForeground(Qt.white)
                elif load_sheet.cell(i, j).fill.start_color.index == 'FF008000':
                    self.table.item(i - 1, j - 1).setBackground(Qt.darkGreen)
                    self.table.item(i - 1, j - 1).setForeground(Qt.white)
                elif load_sheet.cell(i, j).fill.start_color.index == 'FFFF0000':
                    self.table.item(i - 1, j - 1).setBackground(Qt.red)
                    self.table.item(i - 1, j - 1).setForeground(Qt.white)
                elif load_sheet.cell(i, j).fill.start_color.index == 'FF808080':
                    self.table.item(i - 1, j - 1).setBackground(Qt.darkGray)
                    self.table.item(i - 1, j - 1).setForeground(Qt.white)
                else:
                    self.table.item(i - 1, j - 1).setForeground(Qt.black)

        edit.clicked.connect(self.btn_edit)
        # self.show() #파일 선택후 맵미리보기창이 뒤에 뜨게됨

    def btn_edit(self):
        self.hide()
        self.fifth = sixthwindow()
        self.fifth.exec()


# 6. editFile.ui
class sixthwindow(QDialog, QWidget, form_sixthwindow):
    def __init__(self, parent=None):
        global temp_count_len, temp_count_wid, row, col, file_name, file_grid
        temp_count_len = int(col)
        temp_count_wid = int(row)
        sql = "SELECT * FROM grid " + "WHERE Grid_ID = %s;"
        cur.execute(sql, [str(file_name)])
        file_grid = cur.fetchone()

        super(sixthwindow, self).__init__()
        # self.initUi()
        # self.setupUi(self)
        self.setWindowTitle("맵 수정하기")
        self.setFixedSize(1000, 550)
        self.table = QTableWidget(parent)
        vbox = QVBoxLayout()
        vbox.addWidget(self.table)
        grid = QGridLayout()
        vbox.addLayout(grid)
        charge = QPushButton("충전")
        grid.addWidget(charge, 0, 0)
        chute = QPushButton("슈트")
        grid.addWidget(chute, 0, 1)
        ws = QPushButton("워크스테이션")
        grid.addWidget(ws, 0, 2)
        buffer = QPushButton("버퍼")
        grid.addWidget(buffer, 1, 0)
        block = QPushButton("블락")
        grid.addWidget(block, 1, 1)
        trash = QPushButton("삭제")
        grid.addWidget(trash, 0, 9)
        clear = QPushButton("초기화")
        grid.addWidget(clear, 1, 9)
        addrow = QPushButton("row추가")
        grid.addWidget(addrow, 0, 7)
        addcol = QPushButton("col추가")
        grid.addWidget(addcol, 0, 8)
        delrow = QPushButton("row삭제")
        grid.addWidget(delrow, 1, 7)
        delcol = QPushButton("col삭제")
        grid.addWidget(delcol, 1, 8)
        save = QPushButton("저장")
        grid.addWidget(save, 2, 10)
        north = but_img(u_path)
        grid.addWidget(north, 0, 4, 1, 2)
        south = but_img(d_path)
        grid.addWidget(south, 2, 4, 1, 2)
        west = but_img(l_path)
        grid.addWidget(west, 1, 3)
        east = but_img(r_path)
        grid.addWidget(east, 1, 6)
        u_r = but_img(u_r_path)
        grid.addWidget(u_r, 0, 6)
        d_r = but_img(d_r_path)
        grid.addWidget(d_r, 2, 6)
        l_d = but_img(d_l_path)
        grid.addWidget(l_d, 2, 3)
        l_u = but_img(u_l_path)
        grid.addWidget(l_u, 0, 3)
        l_r = but_img(r_l_path)
        grid.addWidget(l_r, 1, 5)
        u_d=but_img(u_d_path)
        grid.addWidget(u_d, 1, 4)
        self.setLayout(vbox)
        self.setGeometry(200, 200, 400, 500)
        load_xlsx = openpyxl.load_workbook(filename, data_only=True)
        load_sheet = load_xlsx['NewSheet1']
        # row = load_sheet.max_row
        # col = load_sheet.max_column
        self.table.setColumnCount(col)
        self.table.setRowCount(row)
        # 반드시 item 생성해야 셀 색상 변경가능
        for i in range(row):
            for j in range(col):
                self.table.setItem(i, j, QTableWidgetItem())
        self.table.resizeColumnsToContents()
        self.table.resizeRowsToContents()
        # load_excel은 1부터, table은 0부터
        cnum = 1
        for i in range(1, row + 1):
            for j in range(1, col + 1):
                cell_num = str(file_name) + '_c' + str(cnum).zfill(4)
                cnum = cnum + 1
                sql = "SELECT * FROM cell " + "WHERE Cell_ID = %s;"
                cur.execute(sql, [cell_num])
                cellinfo = cur.fetchone()
                # n:6 s:7 w:8 e:9
                if cellinfo[6] == 1:
                    item = dir_img(u_path, "↑")
                    self.table.setItem(i - 1, j - 1, item)
                    # self.table.item(i - 1, j - 1).setText("↑")
                elif cellinfo[7] == 1:
                    item = dir_img(d_path, "↓")
                    self.table.setItem(i - 1, j - 1, item)
                    # self.table.item(i - 1, j - 1).setText("↓")
                elif cellinfo[8] == 1:
                    item = dir_img(l_path, "←")
                    self.table.setItem(i - 1, j - 1, item)
                    # self.table.item(i - 1, j - 1).setText("←")
                elif cellinfo[9] == 1:
                    item = dir_img(r_path, "→")
                    self.table.setItem(i - 1, j - 1, item)
                    # self.table.item(i - 1, j - 1).setText("→")
                """elif load_sheet.cell(i, j).text()=="←↓": #db로 바꾸기
                    item = dir_img(d_l_path, "←↓")
                    self.table.setItem(i - 1, j - 1, item)
                elif load_sheet.cell(i, j).text()=="↓→": #db로 바꾸기
                    item = dir_img(d_r_path, "↓→")
                    self.table.setItem(i - 1, j - 1, item)
                elif load_sheet.cell(i, j).text()=="↔": #db로 바꾸기
                    item = dir_img(r_l_path, "↔")
                    self.table.setItem(i - 1, j - 1, item)
                elif load_sheet.cell(i, j).text()=="↕": #db로 바꾸기
                    item = dir_img(u_d_path, "↕")
                    self.table.setItem(i - 1, j - 1, item)
                elif load_sheet.cell(i, j).text()=="←↑": #db로 바꾸기
                    item = dir_img(u_l_path, "←↑")
                    self.table.setItem(i - 1, j - 1, item)
                elif load_sheet.cell(i, j).text()=="↑→": #db로 바꾸기
                    item = dir_img(u_r_path, "↑→")
                    self.table.setItem(i - 1, j - 1, item)"""
                if load_sheet.cell(i, j).fill.start_color.index == 'FFFFFF00':
                    self.table.item(i - 1, j - 1).setBackground(Qt.yellow)
                    self.table.item(i - 1, j - 1).setForeground(Qt.black)
                elif load_sheet.cell(i, j).fill.start_color.index == 'FF0000FF':
                    self.table.item(i - 1, j - 1).setBackground(Qt.darkBlue)
                    self.table.item(i - 1, j - 1).setForeground(Qt.white)
                elif load_sheet.cell(i, j).fill.start_color.index == 'FF008000':
                    self.table.item(i - 1, j - 1).setBackground(Qt.darkGreen)
                    self.table.item(i - 1, j - 1).setForeground(Qt.white)
                elif load_sheet.cell(i, j).fill.start_color.index == 'FFFF0000':
                    self.table.item(i - 1, j - 1).setBackground(Qt.red)
                    self.table.item(i - 1, j - 1).setForeground(Qt.white)
                elif load_sheet.cell(i, j).fill.start_color.index == 'FF808080':
                    self.table.item(i - 1, j - 1).setBackground(Qt.darkGray)
                    self.table.item(i - 1, j - 1).setForeground(Qt.white)
                else:
                    self.table.item(i - 1, j - 1).setForeground(Qt.black)

        charge.clicked.connect(self.btn_charge)  # charge button 클릭
        chute.clicked.connect(self.btn_chute)  # chute button 클릭
        ws.clicked.connect(self.btn_ws)  # ws button 클릭
        buffer.clicked.connect(self.btn_buffer)  # buffer button 클릭
        block.clicked.connect(self.btn_block)  # block button 클릭
        trash.clicked.connect(self.btn_trash)  # trash button 클릭
        clear.clicked.connect(self.btn_clear)  # clear button 클릭
        addrow.clicked.connect(self.btn_addrow)  # addRow button 클릭
        addcol.clicked.connect(self.btn_addcol)  # addCol button 클릭
        delrow.clicked.connect(self.btn_delrow)  # delRow button 클릭
        delcol.clicked.connect(self.btn_delcol)  # delCol button 클릭
        save.clicked.connect(self.btn_save_map)  # saveMap button 클릭
        north.clicked.connect(self.btn_north)
        south.clicked.connect(self.btn_south)
        west.clicked.connect(self.btn_west)
        east.clicked.connect(self.btn_east)
        u_r.clicked.connect(self.btn_u_r)
        d_r.clicked.connect(self.btn_d_r)
        l_d.clicked.connect(self.btn_l_d)
        l_u.clicked.connect(self.btn_l_u)
        l_r.clicked.connect(self.btn_l_r)
        u_d.clicked.connect(self.btn_u_d)

        self.show()

    def btn_north(self):
        for ix in self.table.selectedIndexes():
            """if self.table.item(ix.row(),ix.column()).background().color()==Qt.white:
                self.table.item(ix.row(), ix.column()).setForeground(Qt.black)
            elif self.table.item(ix.row(),ix.column()).background().color()==Qt.yellow:
                self.table.item(ix.row(), ix.column()).setForeground(Qt.black)
            else:
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)"""
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(u_path, "↑")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
            # self.table.item(ix.row(),ix.column()).setText("↑")

    def btn_south(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(d_path, "↓")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
            # self.table.item(ix.row(),ix.column()).setText("↓")

    def btn_west(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(l_path, "←")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
            # self.table.item(ix.row(),ix.column()).setText("←")

    def btn_east(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(r_path, "→")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
            # self.table.item(ix.row(),ix.column()).setText("→")
    def btn_u_r(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(u_r_path, "↑→")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
    def btn_d_r(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(d_r_path, "↓→")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
    def btn_l_d(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(d_l_path, "←↓")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
    def btn_l_u(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(u_l_path, "←↑")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
    def btn_l_r(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(r_l_path, "↔")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
    def btn_u_d(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(u_d_path, "↕")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)

    def btn_charge(self):
        global yellow, red, green, blue, gray, file_grid
        i_charge = file_grid[13]
        for ix in self.table.selectedIndexes():
            if i_charge == 1:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.yellow)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.black)
                yellow = 1
            elif i_charge == 2:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.red)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                red = 1
            elif i_charge == 3:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGreen)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                green = 1
            elif i_charge == 4:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkBlue)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.shite)
                blue = 1
            else:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGray)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                gray = 1

    def btn_chute(self):
        global yellow, red, green, blue, gray, file_grid
        i_chute = file_grid[14]
        for ix in self.table.selectedIndexes():
            if i_chute == 1:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.yellow)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.black)
                yellow = 2
            elif i_chute == 2:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.red)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                red = 2
            elif i_chute == 3:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGreen)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                green = 2
            elif i_chute == 4:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkBlue)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                blue = 3
            else:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGray)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                gray = 2

    def btn_ws(self):
        global yellow, red, green, blue, gray, file_grid
        i_ws = file_grid[15]
        for ix in self.table.selectedIndexes():
            if i_ws == 1:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.yellow)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.black)
                yellow = 3
            elif i_ws == 2:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.red)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                red = 3
            elif i_ws == 3:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGreen)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                green = 3
            elif i_ws == 4:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkBlue)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                blue = 3
            else:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGray)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                gray = 3

    def btn_buffer(self):
        global yellow, red, green, blue, gray, file_grid
        i_buf = file_grid[16]
        for ix in self.table.selectedIndexes():
            if i_buf == 1:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.yellow)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.black)
                yellow = 4
            elif i_buf == 2:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.red)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                red = 4
            elif i_buf == 3:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGreen)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                green = 4
            elif i_buf == 4:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkBlue)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                blue = 4
            else:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGray)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                gray = 4

    def btn_block(self):
        global yellow, red, green, blue, gray, file_grid
        i_blk = file_grid[17]
        for ix in self.table.selectedIndexes():
            if i_blk == 1:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.yellow)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.black)
                yellow = 5
            elif i_blk == 2:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.red)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                red = 5
            elif i_blk == 3:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGreen)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                green = 5
            elif i_blk == 4:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkBlue)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                blue = 5
            else:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGray)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                gray = 5

    def btn_trash(self):
        for ix in self.table.selectedIndexes():
            item = self.table.item(ix.row(), ix.column())
            item = dir_img(zero_path, "")
            item.setBackground(Qt.white)
            item.setText("")
            item.setForeground(Qt.black)
            self.table.setItem(ix.row(), ix.column(), item)

    def btn_clear(self):
        # 색상 변경 위한 item 추가
        row_count = self.table.rowCount()
        col_count = self.table.columnCount()
        for i in range(row_count):
            for j in range(col_count):
                item = self.table.item(i, j)
                item = dir_img(zero_path, "")
                item.setBackground(Qt.white)
                item.setText("")
                item.setForeground(Qt.black)
                self.table.setItem(i, j, item)

    def btn_addcol(self):
        global temp_count_len
        temp_count_len += 1
        row_count = self.table.rowCount()
        col_count = self.table.columnCount()
        self.table.insertColumn(col_count)  # 새로운 행 count
        # 셀 색상 변경 위해 item 추가
        for i in range(row_count):
            self.table.setItem(i, col_count, QTableWidgetItem())
            self.table.item(i, col_count).setForeground(Qt.black)

    def btn_addrow(self):
        global temp_count_wid
        temp_count_wid += 1
        row_count = self.table.rowCount()
        col_count = self.table.columnCount()
        self.table.insertRow(row_count)
        # 셀 색상 변경 위해 item 추가
        for j in range(col_count):
            self.table.setItem(row_count, j, QTableWidgetItem())
            self.table.item(row_count, j).setForeground(Qt.black)

    def btn_delcol(self):
        global temp_count_len
        temp_count_len -= 1
        col_count = self.table.columnCount()
        self.table.removeColumn(col_count - 1)

    def btn_delrow(self):
        global temp_count_wid
        temp_count_wid -= 1
        row_count = self.table.rowCount()
        self.table.removeRow(row_count - 1)

    # -saveMap button 함수: 맵 저장
    def btn_save_map(self):
        workbook = xlsxwriter.Workbook(filename)  # 지정 파일 이름
        worksheet1 = workbook.add_worksheet('NewSheet1')
        global yellow, red, blue, gray, green, temp_count_len, temp_count_wid, file_name, file_grid
        sql = "CALL deleteGrid(%s); CALL createGrid(%s, %s, %s, %s, %s); CALL updateCellCnt(%s, %s, %s, %s, %s, %s); CALL updateGridColor(%s, %s, %s, %s, %s, %s);"
        cur.execute(sql, [file_name, file_name, temp_count_len, temp_count_wid, int(file_grid[4]), int(file_grid[5]),
                          file_name, int(file_grid[7]), int(file_grid[8]), int(file_grid[9]), int(file_grid[10]),
                          int(file_grid[11]), file_name, int(file_grid[13]), int(file_grid[14]), int(file_grid[15]),
                          int(file_grid[16]), int(file_grid[17])])

        for i in range(13, 18):
            if int(file_grid[i]) == 1:
                yellow = i - 12
            elif int(file_grid[i]) == 2:
                red = i - 12
            elif int(file_grid[i]) == 3:
                green = i - 12
            elif int(file_grid[i]) == 4:
                blue = i - 12
            else:
                gray = i - 12

        cnum = 1
        CSnum = 1
        CHnum = 1
        WSnum = 1
        BUFnum = 1

        for row in range(self.table.rowCount()):
            # rowData=[]
            for col in range(self.table.columnCount()):
                cell_num = str(file_name) + '_c' + str(cnum).zfill(4)
                cnum += 1
                # 셀 생성
                sql = "CALL createCell(%s, %s, %s, %s);"
                cur.execute(sql, [file_name, cell_num, str(row), str(col)])
                item = self.table.item(row, col)
                # worksheet1.write(row, col, item.text())
                format = workbook.add_format()

                # DB에 특수 셀 색상 정보 업데이트
                if item.background().color() == Qt.yellow:
                    sql = "CALL updateCellStatus(%s, %s, %s);"
                    cur.execute(sql, [str(file_name), cell_num, str(yellow)])
                    format.set_bg_color('yellow')
                    format.set_font_color('black')
                elif item.background().color() == Qt.darkBlue:
                    sql = "CALL updateCellStatus(%s, %s, %s);"
                    cur.execute(sql, [str(file_name), cell_num, str(blue)])
                    format.set_bg_color('blue')
                    format.set_font_color('white')
                elif item.background().color() == Qt.darkGreen:
                    sql = "CALL updateCellStatus(%s, %s, %s);"
                    cur.execute(sql, [str(file_name), cell_num, str(green)])
                    format.set_bg_color('green')
                    format.set_font_color('white')
                elif item.background().color() == Qt.red:
                    sql = "CALL updateCellStatus(%s, %s, %s);"
                    cur.execute(sql, [str(file_name), cell_num, str(red)])
                    format.set_bg_color('red')
                    format.set_font_color('white')
                elif item.background().color() == Qt.darkGray:
                    sql = "CALL updateCellStatus(%s, %s, %s);"
                    cur.execute(sql, [str(file_name), cell_num, str(gray)])
                    format.set_bg_color('gray')
                    format.set_font_color('white')
                else:  # 색상 없는 셀도 방향 정보 생김
                    format.set_bg_color('white')
                    format.set_font_color('black')
                if item.text() == "↑":
                    worksheet1.write(row, col, "↑", format)
                    sql = "CALL updateCellDirection(%s, 1, 0, 0, 0);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "↓":
                    worksheet1.write(row, col, "↓", format)
                    sql = "CALL updateCellDirection(%s, 0, 1, 0, 0);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "←":
                    worksheet1.write(row, col, "←", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 1, 0);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "→":
                    worksheet1.write(row, col, "→", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 0, 1);"
                    cur.execute(sql, [cell_num])
                else:
                    worksheet1.write(row, col, "", format)
                """디비 추가
                elif item.text() == "↑→":
                    worksheet1.write(row, col, "↑→", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 0, 1);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "↓→":
                    worksheet1.write(row, col, "↓→", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 0, 1);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "←↓":
                    worksheet1.write(row, col, "←↓", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 0, 1);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "←↑":
                    worksheet1.write(row, col, "←↑", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 0, 1);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "↔":
                    worksheet1.write(row, col, "↔", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 0, 1);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "↕":
                    worksheet1.write(row, col, "↕", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 0, 1);"
                    cur.execute(sql, [cell_num])"""

                sql = "SELECT CellStatus FROM cell " + "WHERE Cell_ID = %s"
                cur.execute(sql, [cell_num])
                Cstatus = cur.fetchone()
                if Cstatus[0] == 1:  # 충전
                    CS_num = str(file_name) + '_CS' + str(CSnum).zfill(4)
                    CSnum += 1
                    sql = "CALL createCS(%s, %s, %s, NULL);"
                    cur.execute(sql, [file_name, cell_num, CS_num])
                elif Cstatus[0] == 2:  # 슈트
                    CH_num = str(file_name) + '_CH' + str(CHnum).zfill(4)
                    CHnum += 1
                    sql = "CALL createChute(%s, %s, %s, NULL, NULL);"
                    cur.execute(sql, [file_name, cell_num, CH_num])
                elif Cstatus[0] == 3:  # 워크스테이션
                    WS_num = str(file_name) + '_WS' + str(WSnum).zfill(4)
                    WSnum += 1
                    sql = "CALL createWS(%s, %s, %s, NULL);"
                    cur.execute(sql, [file_name, cell_num, WS_num])
                elif Cstatus[0] == 4:  # 버퍼
                    BUF_num = str(file_name) + '_BUF' + str(BUFnum).zfill(4)
                    BUFnum += 1
                    sql = "CALL createBuffer(%s, %s, %s);"
                    cur.execute(sql, [file_name, cell_num, BUF_num])
                # 색상 없는 셀 도 방향 정보 추가 해야합니다
        workbook.close()
        self.close()


# 2.setGrid.ui
class secondwindow(QDialog, QWidget, form_secondwindow):
    def __init__(self):
        super(secondwindow, self).__init__()
        # self.initUi()
        self.setupUi(self)
        self.setWindowTitle("새 파일 만들기 - 그리드 설정")
        self.setFixedSize(1000, 550)
        self.show()
        self.cb1.activated[int].connect(self.selectone)
        self.cb2.activated[int].connect(self.selecttwo)
        self.cb3.activated[int].connect(self.selectthr)
        self.cb4.activated[int].connect(self.selectfour)
        self.gridNext.clicked.connect(self.btn_next_to_setattribute)  # gridNext button 클릭

    def selectone(self, int):
        global size_len
        size_len = int

    def selecttwo(self, int):
        global size_wid
        size_wid = int

    def selectthr(self, int):
        global count_len
        count_len = int

    def selectfour(self, int):
        global count_wid
        count_wid = int

    # -gridNext button 함수: setAttribute.ui로 창전환, DB저장
    def btn_next_to_setattribute(self, text):
        global size_len, size_wid, count_len, count_wid
        self.hide()
        self.second = thirdwindow()
        self.second.exec()
        # self.show()


# 3.setAttribute.ui
class thirdwindow(QDialog, QWidget, form_thirdwindow):
    def __init__(self):
        super(thirdwindow, self).__init__()
        # self.initUi()
        self.setupUi(self)
        self.setWindowTitle("새 파일 만들기 - 셀 설정")
        self.setFixedSize(1000, 550)
        global i_charge, i_chute, i_ws, i_buf, i_blk
        i_charge = 0
        i_chute = 0
        i_ws = 0
        i_buf = 0
        i_blk = 0
        self.cb1.activated[int].connect(self.selectone)
        self.cb2.activated[int].connect(self.selecttwo)
        self.cb3.activated[int].connect(self.selectthr)
        self.cb4.activated[int].connect(self.selectfour)
        self.cb5.activated[int].connect(self.selectfive)
        self.btn1.clicked.connect(self.btn_charge_color)
        self.btn2.clicked.connect(self.btn_chute_color)
        self.btn3.clicked.connect(self.btn_ws_color)
        self.btn4.clicked.connect(self.btn_buf_color)
        self.btn5.clicked.connect(self.btn_blk_color)
        self.attributeNext.clicked.connect(self.btn_next_to_map)  # attributeNext button 클릭
        self.show()

    def selectone(self, int):
        global count_charge
        count_charge = int

    def selecttwo(self, int):
        global count_chute
        count_chute = int

    def selectthr(self, int):
        global count_ws
        count_ws = int

    def selectfour(self, int):
        global count_buf
        count_buf = int

    def selectfive(self, int):
        global count_blk
        count_blk = int

    def btn_charge_color(self):
        # 버튼 클릭시 색상 변경 위한 변수(여러 색상)
        global i_charge
        global color_charge
        if i_charge == 5:
            i_charge = 0
        if i_charge == 0:
            self.btn1.setStyleSheet('background:yellow')
            color_charge = "yellow"
        if i_charge == 1:
            self.btn1.setStyleSheet('background:red')
            color_charge = "red"
        if i_charge == 2:
            self.btn1.setStyleSheet('background:green')
            color_charge = "green"
        if i_charge == 3:
            self.btn1.setStyleSheet('background:blue')
            color_charge = "blue"
        if i_charge == 4:
            self.btn1.setStyleSheet('background:darkGray')
            color_charge = "darkGray"
        i_charge = i_charge + 1

    def btn_chute_color(self):
        # 버튼 클릭시 색상 변경 위한 변수(여러 색상)
        global i_chute
        global color_chute
        if i_chute == 5:
            i_chute = 0
        if i_chute == 0:
            self.btn2.setStyleSheet('background:yellow')
            color_chute = "yellow"
        if i_chute == 1:
            self.btn2.setStyleSheet('background:red')
            color_chute = "red"
        if i_chute == 2:
            self.btn2.setStyleSheet('background:green')
            color_chute = "green"
        if i_chute == 3:
            self.btn2.setStyleSheet('background:blue')
            color_chute = "blue"
        if i_chute == 4:
            self.btn2.setStyleSheet('background:darkGray')
            color_chute = "darkGray"
        i_chute = i_chute + 1

    def btn_ws_color(self):
        # 버튼 클릭시 색상 변경 위한 변수(여러 색상)
        global i_ws
        global color_ws
        if i_ws == 5:
            i_ws = 0
        if i_ws == 0:
            self.btn3.setStyleSheet('background:yellow')
            color_ws = "yellow"
        if i_ws == 1:
            self.btn3.setStyleSheet('background:red')
            color_ws = "red"
        if i_ws == 2:
            self.btn3.setStyleSheet('background:green')
            color_ws = "green"
        if i_ws == 3:
            self.btn3.setStyleSheet('background:blue')
            color_ws = "blue"
        if i_ws == 4:
            self.btn3.setStyleSheet('background:darkGray')
            color_ws = "darkGray"
        i_ws = i_ws + 1

    def btn_buf_color(self):
        # 버튼 클릭시 색상 변경 위한 변수(여러 색상)
        global i_buf
        global color_buf
        if i_buf == 5:
            i_buf = 0
        if i_buf == 0:
            self.btn4.setStyleSheet('background:yellow')
            color_buf = "yellow"
        if i_buf == 1:
            self.btn4.setStyleSheet('background:red')
            color_buf = "red"
        if i_buf == 2:
            self.btn4.setStyleSheet('background:green')
            color_buf = "green"
        if i_buf == 3:
            self.btn4.setStyleSheet('background:blue')
            color_buf = "blue"
        if i_buf == 4:
            self.btn4.setStyleSheet('background:darkGray')
            color_buf = "darkGray"
        i_buf = i_buf + 1

    def btn_blk_color(self):
        # 버튼 클릭시 색상 변경 위한 변수(여러 색상)
        global i_blk
        global color_blk
        if i_blk == 5:
            i_blk = 0
        if i_blk == 0:
            self.btn5.setStyleSheet('background:yellow')
            color_blk = "yellow"
        if i_blk == 1:
            self.btn5.setStyleSheet('background:red')
            color_blk = "red"
        if i_blk == 2:
            self.btn5.setStyleSheet('background:green')
            color_blk = "green"
        if i_blk == 3:
            self.btn5.setStyleSheet('background:blue')
            color_blk = "blue"
        if i_blk == 4:
            self.btn5.setStyleSheet('background:darkGray')
            color_blk = "darkGray"
        i_blk = i_blk + 1

    # -attributeNext button 함수: createMap.ui로 창전환
    def btn_next_to_map(self):
        self.hide()
        self.third = fourthwindow()
        self.third.exec()


# 4.createMap.ui
class fourthwindow(QDialog, QWidget, form_fourthwindow):
    def __init__(self, parent=None):
        global temp_count_len, temp_count_wid, count_len, count_wid
        temp_count_len = int(count_len)
        temp_count_wid = int(count_wid)
        super(fourthwindow, self).__init__()
        # self.initUi()
        # self.setupUi(self)
        self.setWindowTitle("새 파일 만들기 - 맵 그리기")
        self.setFixedSize(1000, 550)
        self.table = QTableWidget(parent)
        vbox = QVBoxLayout()
        vbox.addWidget(self.table)
        grid = QGridLayout()
        vbox.addLayout(grid)
        charge = QPushButton("충전")
        grid.addWidget(charge, 0, 0)
        chute = QPushButton("슈트")
        grid.addWidget(chute, 0, 1)
        ws = QPushButton("워크스테이션")
        grid.addWidget(ws, 0, 2)
        buffer = QPushButton("버퍼")
        grid.addWidget(buffer, 1, 0)
        block = QPushButton("블락")
        grid.addWidget(block, 1, 1)
        trash = QPushButton("삭제")
        grid.addWidget(trash, 0, 9)
        clear = QPushButton("초기화")
        grid.addWidget(clear, 1, 9)
        addrow = QPushButton("row추가")
        grid.addWidget(addrow, 0, 7)
        addcol = QPushButton("col추가")
        grid.addWidget(addcol, 0, 8)
        delrow = QPushButton("row삭제")
        grid.addWidget(delrow, 1, 7)
        delcol = QPushButton("col삭제")
        grid.addWidget(delcol, 1, 8)
        save = QPushButton("저장")
        grid.addWidget(save, 2, 10)
        north = but_img(u_path)
        grid.addWidget(north, 0, 4, 1, 2)
        south = but_img(d_path)
        grid.addWidget(south, 2, 4, 1, 2)
        west = but_img(l_path)
        grid.addWidget(west, 1, 3)
        east = but_img(r_path)
        grid.addWidget(east, 1, 6)
        u_r = but_img(u_r_path)
        grid.addWidget(u_r, 0, 6)
        d_r = but_img(d_r_path)
        grid.addWidget(d_r, 2, 6)
        l_d = but_img(d_l_path)
        grid.addWidget(l_d, 2, 3)
        l_u = but_img(u_l_path)
        grid.addWidget(l_u, 0, 3)
        l_r = but_img(r_l_path)
        grid.addWidget(l_r, 1, 5)
        u_d = but_img(u_d_path)
        grid.addWidget(u_d, 1, 4)
        self.setLayout(vbox)
        self.setGeometry(200, 200, 400, 500)
        self.table.setColumnCount(int(count_len))
        self.table.setRowCount(int(count_wid))
        for i in range(int(count_wid)):
            for j in range(int(count_len)):
                self.table.setItem(i, j, QTableWidgetItem())
                # self.table.item(i , j).setBackground(Qt.white)
                # self.table.item(i, j).setForeground(Qt.black)
        self.table.resizeColumnsToContents()
        self.table.resizeRowsToContents()
        charge.clicked.connect(self.btn_charge)  # charge button 클릭
        chute.clicked.connect(self.btn_chute)  # chute button 클릭
        ws.clicked.connect(self.btn_ws)  # ws button 클릭
        buffer.clicked.connect(self.btn_buffer)  # buffer button 클릭
        block.clicked.connect(self.btn_block)  # block button 클릭
        trash.clicked.connect(self.btn_trash)  # trash button 클릭
        clear.clicked.connect(self.btn_clear)  # clear button 클릭
        addrow.clicked.connect(self.btn_addrow)  # addRow button 클릭
        addcol.clicked.connect(self.btn_addcol)  # addCol button 클릭
        delrow.clicked.connect(self.btn_delrow)  # delRow button 클릭
        delcol.clicked.connect(self.btn_delcol)  # delCol button 클릭
        save.clicked.connect(self.btn_save_map)  # saveMap button 클릭
        north.clicked.connect(self.btn_north)
        south.clicked.connect(self.btn_south)
        west.clicked.connect(self.btn_west)
        east.clicked.connect(self.btn_east)
        u_r.clicked.connect(self.btn_u_r)
        d_r.clicked.connect(self.btn_d_r)
        l_d.clicked.connect(self.btn_l_d)
        l_u.clicked.connect(self.btn_l_u)
        l_r.clicked.connect(self.btn_l_r)
        u_d.clicked.connect(self.btn_u_d)
        self.show()

    def btn_north(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(u_path, "↑")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
            # self.table.item(ix.row(),ix.column()).setText("↑")

    def btn_south(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(d_path, "↓")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
            # self.table.item(ix.row(),ix.column()).setText("↓")

    def btn_west(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(l_path, "←")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
            # self.table.item(ix.row(),ix.column()).setText("←")

    def btn_east(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(r_path, "→")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
            # self.table.item(ix.row(),ix.column()).setText("→")
    def btn_u_r(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(u_r_path, "↑→")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
    def btn_d_r(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(d_r_path, "↓→")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
    def btn_l_d(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(d_l_path, "←↓")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
    def btn_l_u(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(u_l_path, "←↑")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
    def btn_l_r(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(r_l_path, "↔")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
    def btn_u_d(self):
        for ix in self.table.selectedIndexes():
            bgcolor = self.table.item(
                ix.row(), ix.column()).background().color()
            item = dir_img(u_d_path, "↕")
            self.table.setItem(ix.row(), ix.column(), item)
            if bgcolor != QColor.fromRgbF(0, 0, 0, 1):
                self.table.item(ix.row(), ix.column()).setBackground(bgcolor)
    # @pyqtSlot()
    def btn_charge(self):
        global yellow, red, green, blue, gray
        for ix in self.table.selectedIndexes():
            if i_charge == 1:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.yellow)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.black)
                yellow = 1
            elif i_charge == 2:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.red)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                red = 1
            elif i_charge == 3:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGreen)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                green = 1
            elif i_charge == 4:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkBlue)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                blue = 1
            else:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGray)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                gray = 1

    def btn_chute(self):
        global yellow, red, green, blue, gray
        for ix in self.table.selectedIndexes():
            if i_chute == 1:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.yellow)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.black)
                yellow = 2
            elif i_chute == 2:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.red)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                red = 2
            elif i_chute == 3:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGreen)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                green = 2
            elif i_chute == 4:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkBlue)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                blue = 3
            else:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGray)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                gray = 2

    def btn_ws(self):
        global yellow, red, green, blue, gray
        for ix in self.table.selectedIndexes():
            if i_ws == 1:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.yellow)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.black)
                yellow = 3
            elif i_ws == 2:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.red)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                red = 3
            elif i_ws == 3:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGreen)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                green = 3
            elif i_ws == 4:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkBlue)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                blue = 3
            else:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGray)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                gray = 3

    def btn_buffer(self):
        global yellow, red, green, blue, gray
        for ix in self.table.selectedIndexes():
            if i_buf == 1:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.yellow)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.black)
                yellow = 4
            elif i_buf == 2:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.red)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                red = 4
            elif i_buf == 3:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGreen)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                green = 4
            elif i_buf == 4:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkBlue)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                blue = 4
            else:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGray)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                gray = 4

    def btn_block(self):
        global yellow, red, green, blue, gray
        for ix in self.table.selectedIndexes():
            if i_blk == 1:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.yellow)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.black)
                yellow = 5
            elif i_blk == 2:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.red)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                red = 5
            elif i_blk == 3:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGreen)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                green = 5
            elif i_blk == 4:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkBlue)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                blue = 5
            else:
                self.table.item(ix.row(), ix.column()).setBackground(Qt.darkGray)
                self.table.item(ix.row(), ix.column()).setForeground(Qt.white)
                gray = 5

    def btn_trash(self):
        for ix in self.table.selectedIndexes():
            item=self.table.item(ix.row(), ix.column())
            item = dir_img(zero_path, "")
            item.setBackground(Qt.white)
            item.setText("")
            item.setForeground(Qt.black)
            self.table.setItem(ix.row(), ix.column(), item)

    def btn_clear(self):
        row_count = self.table.rowCount()
        col_count = self.table.columnCount()
        for i in range(row_count):
            for j in range(col_count):
                item = self.table.item(i,j)
                item = dir_img(zero_path, "")
                item.setBackground(Qt.white)
                item.setText("")
                item.setForeground(Qt.black)
                self.table.setItem(i, j, item)

    def btn_addcol(self):
        global temp_count_len
        temp_count_len += 1
        row_count = self.table.rowCount()
        col_count = self.table.columnCount()
        self.table.insertColumn(col_count)  # 새로운 행 count
        for i in range(row_count):
            self.table.setItem(i, col_count, QTableWidgetItem())
            self.table.item(i, col_count).setBackground(Qt.white)
            self.table.item(i, col_count).setForeground(Qt.black)

    def btn_addrow(self):
        global temp_count_wid
        temp_count_wid += 1
        row_count = self.table.rowCount()
        col_count = self.table.columnCount()
        self.table.insertRow(row_count)
        for j in range(col_count):
            self.table.setItem(row_count, j, QTableWidgetItem())
            self.table.item(row_count, j).setBackground(Qt.white)
            self.table.item(row_count, j).setForeground(Qt.black)

    def btn_delcol(self):
        global temp_count_len
        temp_count_len -= 1
        col_count = self.table.columnCount()
        self.table.removeColumn(col_count - 1)

    def btn_delrow(self):
        global temp_count_wid
        temp_count_wid -= 1
        row_count = self.table.rowCount()
        self.table.removeRow(row_count - 1)

    # -saveMap button 함수: 맵 저장
    def btn_save_map(self):
        global yellow, red, blue, gray, green
        global temp_count_len, temp_count_wid, size_len, size_wid  # 그리드 크기, 셀 크기
        global count_charge, count_chute, count_ws, count_buf, count_blk  # 특수 셀 개수
        global i_charge, i_chute, i_ws, i_buf, i_blk  # 특수 셀 색

        file = QFileDialog.getSaveFileName(self, '', '', 'xlsx Files(*.xlsx)')
        workbook = xlsxwriter.Workbook(file[0])  # 지정 파일 이름
        worksheet1 = workbook.add_worksheet('NewSheet1')

        file_name = QFileInfo(file[0]).baseName()

        # 그리드 테이블 생성
        sql = "CALL createGrid(%s, %s, %s, %s, %s);"
        cur.execute(sql, [str(file_name), temp_count_len, temp_count_wid, size_len, size_wid])

        # 그리드 셀 개수, 특수 셀 색상 정보 업데이트
        sql = "CALL updateCellCnt(%s, %s, %s, %s, %s, %s); CALL updateGridColor(%s, %s, %s, %s, %s, %s)"
        cur.execute(sql,
                    [str(file_name), count_charge, count_chute, count_ws, count_buf, count_blk,
                     str(file_name), i_charge, i_chute, i_ws, i_buf, i_blk])

        cnum = 1
        CSnum = 1
        CHnum = 1
        WSnum = 1
        BUFnum = 1

        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                cell_num = str(file_name) + '_c' + str(cnum).zfill(4)
                cnum += 1

                # 기본 셀 생성(형식 : 파일이름_c_num)
                sql = "CALL createCell(%s, %s, %s, %s);"
                cur.execute(sql, [str(file_name), cell_num, str(row), str(col)])

                item = self.table.item(row, col)
                format = workbook.add_format()

                # DB에 특수 셀 색상 정보 업데이트
                if item.background().color() == Qt.yellow:
                    sql = "CALL updateCellStatus(%s, %s, %s);"
                    cur.execute(sql, [str(file_name), cell_num, str(yellow)])
                    format.set_bg_color('yellow')
                    format.set_font_color('black')
                elif item.background().color() == Qt.darkBlue:
                    sql = "CALL updateCellStatus(%s, %s, %s);"
                    cur.execute(sql, [str(file_name), cell_num, str(blue)])
                    format.set_bg_color('blue')
                    format.set_font_color('white')
                elif item.background().color() == Qt.darkGreen:
                    sql = "CALL updateCellStatus(%s, %s, %s);"
                    cur.execute(sql, [str(file_name), cell_num, str(green)])
                    format.set_bg_color('green')
                    format.set_font_color('white')
                elif item.background().color() == Qt.red:
                    sql = "CALL updateCellStatus(%s, %s, %s);"
                    cur.execute(sql, [str(file_name), cell_num, str(red)])
                    format.set_bg_color('red')
                    format.set_font_color('white')
                elif item.background().color() == Qt.darkGray:
                    sql = "CALL updateCellStatus(%s, %s, %s);"
                    cur.execute(sql, [str(file_name), cell_num, str(gray)])
                    format.set_bg_color('gray')
                    format.set_font_color('white')
                else:
                    format.set_bg_color('white')
                    format.set_font_color('black')
                if item.text() == "↑":
                    worksheet1.write(row, col, "↑", format)
                    sql = "CALL updateCellDirection(%s, 1, 0, 0, 0);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "↓":
                    worksheet1.write(row, col, "↓", format)
                    sql = "CALL updateCellDirection(%s, 0, 1, 0, 0);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "←":
                    worksheet1.write(row, col, "←", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 1, 0);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "→":
                    worksheet1.write(row, col, "→", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 0, 1);"
                    cur.execute(sql, [cell_num])
                else:
                    worksheet1.write(row, col, "", format)
                """디비 추가
                elif item.text() == "↑→":
                    worksheet1.write(row, col, "↑→", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 0, 1);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "↓→":
                    worksheet1.write(row, col, "↓→", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 0, 1);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "←↓":
                    worksheet1.write(row, col, "←↓", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 0, 1);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "←↑":
                    worksheet1.write(row, col, "←↑", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 0, 1);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "↔":
                    worksheet1.write(row, col, "↔", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 0, 1);"
                    cur.execute(sql, [cell_num])
                elif item.text() == "↕":
                    worksheet1.write(row, col, "↕", format)
                    sql = "CALL updateCellDirection(%s, 0, 0, 0, 1);"
                    cur.execute(sql, [cell_num])"""



                sql = "SELECT CellStatus FROM cell " + "WHERE Cell_ID = %s;"
                cur.execute(sql, [cell_num])
                Cstatus = cur.fetchone()
                if Cstatus[0] == 1:  # 충전 셀 생성
                    CS_num = str(file_name) + '_CS' + str(CSnum).zfill(4)
                    CSnum += 1
                    sql = "CALL createCS(%s, %s, %s, NULL);"
                    cur.execute(sql, [str(file_name), cell_num, CS_num])
                elif Cstatus[0] == 2:  # 슈트 셀 생성
                    CH_num = str(file_name) + '_CH' + str(CHnum).zfill(4)
                    CHnum += 1
                    sql = "CALL createChute(%s, %s, %s, NULL, NULL);"
                    cur.execute(sql, [str(file_name), cell_num, CH_num])
                elif Cstatus[0] == 3:  # 워크스테이션 셀 생성
                    WS_num = str(file_name) + '_WS' + str(WSnum).zfill(4)
                    WSnum += 1
                    sql = "CALL createWS(%s, %s, %s, NULL);"
                    cur.execute(sql, [str(file_name), cell_num, WS_num])
                elif Cstatus[0] == 4:  # 버퍼 셀 생성
                    BUF_num = str(file_name) + '_BUF' + str(BUFnum).zfill(4)
                    BUFnum += 1
                    sql = "CALL createBuffer(%s, %s, %s);"
                    cur.execute(sql, [str(file_name), cell_num, BUF_num])

        workbook.close()
        self.close()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    myWindow = WindowClass()
    myWindow.show()
    app.exec()

conn.close()
